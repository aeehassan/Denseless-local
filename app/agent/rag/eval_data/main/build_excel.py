"""
build_excel.py
--------------
Reads all JSON eval dataset files from a specified folder and writes them
into a single Excel workbook, one sheet per JSON file.

Algorithm:
  Step 1  — Scan the folder for all .json files
  Step 2  — Parse each JSON file: transform dict-of-dicts into flat rows,
             pulling the outer key as pair_id
  Step 2.5— Render string content: \n -> real newline (Unicode escapes are
             handled automatically by Python's JSON parser on load)
  Step 3  — Build a pandas DataFrame per file
  Step 4  — Write each DataFrame to its own sheet in an Excel workbook,
             with wrap text enabled and column widths auto-adjusted
  Step 5  — Save the workbook to the same folder as eval_dataset.xlsx

Dependencies
# pip install pandas openpyxl
# To run: python eval/build_excel.py
"""

import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
# Folder: app\agent\rag\eval_data\main
FOLDER      = Path(__file__).parent
print(FOLDER)
OUTPUT_FILE = FOLDER / "eval_dataset.xlsx"

# Maximum column width in Excel units — prevents unreadably wide columns
MAX_COL_WIDTH = 80

# Row height in points — accommodates multi-line cell content
ROW_HEIGHT = 80


# ── Step 2.5 Helper — Render string content ───────────────────────────────────
def render_value(value) -> str:
    """
    Prepares a cell value for Excel rendering.

    - Unicode escape sequences (e.g. \u2014 -> —) are already resolved
      automatically by Python's json.load(), so no action needed there.
    - \n escape sequences are converted to real newline characters so Excel
      displays multi-line content within the cell rather than literal \n.
    - All other types are returned as-is.
    """
    if isinstance(value, str):
        # Replace escaped newline sequences with real newlines.
        # json.load() gives us a real \n character in memory (not the two
        # characters backslash-n), so this catches cases where the source
        # data was double-escaped or loaded as raw text.
        return value.replace("\\n", "\n")
    if isinstance(value, list):
        # For list fields (e.g. ltm_memories), join items with real newlines
        # so each memory appears on its own line in the cell.
        return "\n".join(str(item).replace("\\n", "\n") for item in value)
    return value


# ── Step 1 — Scan the folder ──────────────────────────────────────────────────
def collect_json_files(folder: Path) -> list[Path]:
    """
    Returns all .json files in the folder, sorted alphabetically.
    Skips any non-.json files silently.
    """
    files = sorted(folder.glob("*.json"))
    if not files:
        raise FileNotFoundError(f"No JSON files found in: {folder}")
    return files


# ── Step 2 — Parse each JSON file ────────────────────────────────────────────
def parse_json_to_rows(filepath: Path) -> list[dict]:
    """
    Reads a JSON file whose structure is a dict-of-dicts:
      {
        "001": { "field_a": "...", "field_b": "..." },
        "002": { "field_a": "...", "field_b": "..." },
        ...
      }

    Transforms it into a flat list of row dicts, injecting the outer key
    as the first column 'pair_id'. Applies Step 2.5 rendering to all values.
    """
    with open(filepath, encoding="utf-8") as f:
        data: dict = json.load(f)

    rows = []
    for pair_id, fields in data.items():
        # Pull the outer key as pair_id, then spread inner fields as columns.
        row = {"pair_id": pair_id}
        for col, value in fields.items():
            # Step 2.5 — render string content before storing in the row
            row[col] = render_value(value)
        rows.append(row)

    return rows


# ── Step 3 — Build a DataFrame per file ──────────────────────────────────────
def build_dataframe(rows: list[dict]) -> pd.DataFrame:
    """
    Converts a list of row dicts into a pandas DataFrame.
    pair_id is guaranteed to be the first column.
    """
    df = pd.DataFrame(rows)
    # Ensure pair_id is always the leftmost column regardless of dict ordering
    cols = ["pair_id"] + [c for c in df.columns if c != "pair_id"]
    return df[cols]


# ── Step 4 — Write to Excel with formatting ───────────────────────────────────
def write_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str):
    """
    Writes a DataFrame to a named sheet in the Excel workbook.

    Formatting applied:
      - Wrap text enabled on all cells so real newlines render correctly
      - Column widths auto-adjusted based on longest content, capped at
        MAX_COL_WIDTH to prevent unreadably wide columns
      - Fixed row height set to ROW_HEIGHT points to accommodate multi-line
        content without requiring manual row expansion
    """
    df.to_excel(writer, sheet_name=sheet_name, index=False)

    worksheet = writer.sheets[sheet_name]

    # ── Wrap text and auto-width ──────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)

        # Calculate the width needed: max of header length and longest cell
        # value in the column. Use only the first line of multi-line content
        # for width calculation — the full content is handled by row height.
        header_len = len(str(col_name))
        cell_max_len = (
            df[col_name]
            .astype(str)
            .apply(lambda x: max(len(line) for line in x.split("\n")))
            .max()
        )
        col_width = min(max(header_len, cell_max_len) + 4, MAX_COL_WIDTH)
        worksheet.column_dimensions[col_letter].width = col_width

        # Apply wrap text alignment to every data cell in this column
        for row_idx in range(2, len(df) + 2):  # row 1 is the header
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Header alignment ──────────────────────────────────────────────────────
    for col_idx in range(1, len(df.columns) + 1):
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.alignment = Alignment(
            wrap_text=False, vertical="center", horizontal="center"
        )

    # ── Fixed row height for all data rows ────────────────────────────────────
    for row_idx in range(2, len(df) + 2):
        worksheet.row_dimensions[row_idx].height = ROW_HEIGHT


# ── Step 5 — Main entry point ─────────────────────────────────────────────────
def main():
    if not FOLDER.exists():
        raise FileNotFoundError(f"Eval datasets folder not found: {FOLDER}")

    # Step 1 — collect all JSON files in the folder
    json_files = collect_json_files(FOLDER)
    print(f"Found {len(json_files)} JSON file(s):\n")

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for filepath in json_files:
            # Sheet name is the filename without extension
            sheet_name = filepath.stem
            print(f"  Processing: {filepath.name} -> sheet '{sheet_name}'")

            # Step 2 — parse JSON into flat rows (Step 2.5 applied inside)
            rows = parse_json_to_rows(filepath)

            # Step 3 — build DataFrame
            df = build_dataframe(rows)

            # Step 4 — write sheet with formatting
            write_sheet(writer, df, sheet_name)
            print(f"    {len(df)} rows written\n")

    # Step 5 — workbook is saved when the ExcelWriter context exits
    print(f"Excel file saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()